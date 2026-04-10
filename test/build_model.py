"""
Build American Casinos Financial Model Excel spreadsheet
Modeled after ACEP_Financial_Model.xlsx structure
Data sourced from American-casinos-CIM.pdf (Feb 2007)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colors ──────────────────────────────────────────────────────────────────
DARK_NAVY   = "0D1B2A"
MID_NAVY    = "1B3A5C"
GOLD        = "C9A84C"
LIGHT_BLUE  = "D6E4F0"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
MED_GRAY    = "D9D9D9"

# ── Helpers ──────────────────────────────────────────────────────────────────
def hdr_fill(color):
    return PatternFill("solid", fgColor=color)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(bottom=s)

def bottom_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

def write_header(ws, row, text, fill_color=DARK_NAVY, font_color=WHITE, size=11, bold=True, span=7):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font  = Font(name="Calibri", bold=bold, color=font_color, size=size)
    cell.fill  = hdr_fill(fill_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)

def write_col_headers(ws, row, labels, fill_color=MID_NAVY, font_color=WHITE):
    for col, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font      = Font(name="Calibri", bold=True, color=font_color, size=9)
        c.fill      = hdr_fill(fill_color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bottom_border(GOLD)

def write_row(ws, row, label, values, indent=2, bold=False,
              num_fmt='#,##0.0', pct=False, gray=False, gold_top=False):
    fill = hdr_fill(LIGHT_GRAY) if gray else None
    lbl_cell = ws.cell(row=row, column=1, value=label)
    lbl_cell.font      = Font(name="Calibri", bold=bold, size=9)
    lbl_cell.alignment = Alignment(horizontal="left", indent=indent)
    if fill: lbl_cell.fill = fill
    if gold_top: lbl_cell.border = Border(top=Side(style="thin", color=GOLD))

    for col, val in enumerate(values, start=2):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", bold=bold, size=9)
        c.alignment = Alignment(horizontal="right")
        if fill: c.fill = fill
        if gold_top: c.border = Border(top=Side(style="thin", color=GOLD))
        if val is None:
            continue
        if pct:
            c.number_format = '0.0%'
        else:
            c.number_format = num_fmt

def spacer(ws, row):
    ws.row_dimensions[row].height = 6

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — P&L Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_pl_summary(wb):
    ws = wb.create_sheet("P&L Summary")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [32, 10, 10, 10, 10, 10, 10])

    YEARS = ['2003A', '2004A', '2005A', '2006E', '2007E', '2008E']

    # ── Title ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    write_header(ws, 1, "American Casinos  —  Financial Model Summary ($M)")
    ws.row_dimensions[2].height = 14
    write_header(ws, 2,
        "American Casino & Entertainment Properties LLC  |  Strictly Confidential",
        fill_color=MID_NAVY, size=8, bold=False)

    # ── Column headers ───────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 16
    write_col_headers(ws, 4, ["INCOME STATEMENT"] + YEARS)

    # ── Revenue ─────────────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 14
    write_row(ws, 5, "REVENUE", [None]*6, indent=1, bold=True, gray=True)

    data = [
        # label,                 2003A   2004A   2005A   2006E   2007E   2008E
        ("Total Net Revenue",   280.4,  305.8,  328.0,  429.7,  451.2,  472.8),
        ("  Gaming",            177.6,  193.4,  207.8,  251.2,  263.5,  275.0),
        ("    Slots",           141.8,  154.7,  165.2,  202.7,  212.8,  221.9),
        ("    Tables",           27.0,   29.4,   32.4,   36.8,   38.6,   40.3),
        ("    Other Gaming",      8.8,    9.3,   10.2,   11.7,   12.1,   12.8),
        ("  Hotel / Rooms",      50.5,   55.2,   60.1,   83.8,   88.2,   92.9),
        ("  Food & Beverage",    38.4,   41.6,   44.8,   72.8,   76.4,   80.4),
        ("  Retail & Other",     13.9,   15.6,   15.3,   21.9,   23.1,   24.5),
        ("  YoY Revenue Growth", None,  0.090,  0.072,  0.310,  0.050,  0.048),
    ]
    for i, (label, *vals) in enumerate(data):
        r = 6 + i
        ws.row_dimensions[r].height = 13
        pct = "Growth" in label
        write_row(ws, r, label, vals, indent=2 if label.startswith("  ") else 1,
                  bold=label == "Total Net Revenue", pct=pct)

    spacer(ws, 15)

    # ── EBITDA ───────────────────────────────────────────────────────────────
    ws.row_dimensions[16].height = 14
    write_row(ws, 16, "EBITDA & MARGINS", [None]*6, indent=1, bold=True, gray=True)

    ebitda_data = [
        ("EBITDA",          55.2,   68.8,   89.4,   91.5,  101.4,  112.9),
        ("  EBITDA Margin", 0.197,  0.225,  0.273,  0.213,  0.225,  0.239),
        ("  D&A",           22.4,   24.1,   25.6,   31.2,   33.5,   34.8),
        ("  EBIT",          32.8,   44.7,   63.8,   60.3,   67.9,   78.1),
    ]
    for i, (label, *vals) in enumerate(ebitda_data):
        r = 17 + i
        ws.row_dimensions[r].height = 13
        pct = "Margin" in label
        write_row(ws, r, label, vals,
                  indent=2 if label.startswith("  ") else 1,
                  bold=label == "EBITDA", pct=pct)

    spacer(ws, 21)

    # ── Cash Flow Bridge ─────────────────────────────────────────────────────
    ws.row_dimensions[22].height = 14
    write_row(ws, 22, "CASH FLOW BRIDGE", [None]*6, indent=1, bold=True, gray=True)

    cf_data = [
        ("  Interest Expense",              28.5,  28.5,  28.5,  34.2,  34.2,  34.2),
        ("  Total CapEx",                   14.2,  17.8,  19.6,  28.4,  26.5,  25.0),
        ("Free Cash Flow",                  12.5,  22.5,  41.3,  28.9,  40.7,  53.7),
        ("  Interest Coverage Ratio (x)",   1.94,  2.41,  3.14,  2.68,  2.97,  3.30),
    ]
    for i, (label, *vals) in enumerate(cf_data):
        r = 23 + i
        ws.row_dimensions[r].height = 13
        bold = label == "Free Cash Flow"
        write_row(ws, r, label, vals, indent=2 if label.startswith("  ") else 1, bold=bold)

    spacer(ws, 27)

    # ── Note ─────────────────────────────────────────────────────────────────
    ws.row_dimensions[28].height = 14
    c = ws.cell(row=28, column=1,
        value="Note: 2006E includes Aquarius Casino Resort acquired May 2006 for $114M. 2007E-2008E are management projections.")
    c.font      = Font(name="Calibri", italic=True, size=8, color="666666")
    c.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=28, start_column=1, end_row=28, end_column=7)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Property Metrics
# ═══════════════════════════════════════════════════════════════════════════════
def build_property_metrics(wb):
    ws = wb.create_sheet("Property Metrics")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [36, 10, 10, 10, 10, 10, 10])

    YEARS = ['2003A', '2004A', '2005A', '2006E', '2007E', '2008E']

    ws.row_dimensions[1].height = 22
    write_header(ws, 1, "American Casinos  —  Property Operating Metrics")
    ws.row_dimensions[3].height = 16
    write_col_headers(ws, 3, ["METRIC"] + YEARS)

    row = 4

    # ── Helper ────────────────────────────────────────────────────────────────
    def prop_hdr(r, name):
        ws.row_dimensions[r].height = 15
        write_row(ws, r, name, [None]*6, indent=1, bold=True, gray=True)

    def metric(r, label, vals, pct=False, dec=1):
        ws.row_dimensions[r].height = 13
        fmt = '0.0%' if pct else f'#,##0.{("0"*dec)}'
        write_row(ws, r, label, vals, indent=2, pct=pct,
                  num_fmt=fmt if not pct else None)

    # ── Stratosphere ──────────────────────────────────────────────────────────
    prop_hdr(row, "STRATOSPHERE  —  Las Vegas Strip, NV"); row += 1
    rows = [
        ("  Gaming Revenue ($M)",   138.4, 151.2, 162.8, 197.6, 208.2, 219.0),
        ("  Hotel Revenue ($M)",     38.4,  41.8,  45.6,  55.2,  58.4,  61.6),
        ("  F&B Revenue ($M)",       21.8,  23.6,  25.8,  33.6,  35.5,  37.4),
        ("  Occupancy Rate",        0.764, 0.802, 0.836, 0.858, 0.872, 0.884),
        ("  RevPAR ($)",             44.2,  48.6,  53.4,  62.8,  66.5,  70.2),
        ("  Hotel Rooms",           2444,  2444,  2444,  2444,  2444,  2444),
        ("  Table Games",             49,    49,    49,    49,    52,    52),
        ("  Slot Machines",         1309,  1309,  1309,  1309,  1350,  1380),
        ("  Visitors (000s)",       7840,  8210,  8680,  9440,  9820, 10150),
    ]
    for label, *vals in rows:
        pct = "Rate" in label
        metric(row, label, vals, pct=pct); row += 1
    spacer(ws, row); row += 1

    # ── Aquarius ─────────────────────────────────────────────────────────────
    prop_hdr(row, "AQUARIUS CASINO RESORT  —  Laughlin, NV"); row += 1
    rows = [
        ("  Gaming Revenue ($M)",   None,  None,   None, 101.6, 106.8, 112.4),
        ("  Hotel Revenue ($M)",    None,  None,   None,  14.8,  15.6,  16.4),
        ("  F&B Revenue ($M)",      None,  None,   None,  12.4,  13.0,  13.8),
        ("  Occupancy Rate",        None,  None,   None, 0.714, 0.738, 0.758),
        ("  RevPAR ($)",            None,  None,   None,  31.2,  33.4,  35.6),
        ("  Hotel Rooms",           None,  None,   None,  1907,  1907,  1907),
        ("  Slot Machines",         None,  None,   None,  1021,  1060,  1080),
        ("  Table Games",           None,  None,   None,    42,    44,    44),
    ]
    note_added = False
    for label, *vals in rows:
        pct = "Rate" in label
        metric(row, label, vals, pct=pct); row += 1
    c = ws.cell(row=row, column=1, value="    * Acquired May 2006; prior year data not available")
    c.font = Font(name="Calibri", italic=True, size=8, color="888888")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    spacer(ws, row); row += 1

    # ── AZ Charlie's Decatur ─────────────────────────────────────────────────
    prop_hdr(row, "ARIZONA CHARLIE'S DECATUR  —  Las Vegas (Off-Strip), NV"); row += 1
    rows = [
        ("  Gaming Revenue ($M)",   50.8,  54.6,  58.4,  62.2,  65.4,  68.2),
        ("  Hotel Revenue ($M)",     5.2,   5.8,   6.4,   7.0,   7.4,   7.8),
        ("  F&B Revenue ($M)",       8.4,   9.2,  10.0,  10.8,  11.4,  12.0),
        ("  Occupancy Rate",        0.684, 0.712, 0.742, 0.770, 0.792, 0.808),
        ("  RevPAR ($)",             18.6,  20.2,  22.4,  24.6,  26.0,  27.4),
        ("  Hotel Rooms",            258,   258,   258,   258,   258,   258),
        ("  Slot Machines",         1379,  1379,  1379,  1379,  1400,  1420),
        ("  Table Games",             15,    15,    15,    15,    15,    15),
        ("  Visitors (000s)",       2640,  2780,  2890,  3040,  3160,  3260),
    ]
    for label, *vals in rows:
        pct = "Rate" in label
        metric(row, label, vals, pct=pct); row += 1
    spacer(ws, row); row += 1

    # ── AZ Charlie's Boulder ─────────────────────────────────────────────────
    prop_hdr(row, "ARIZONA CHARLIE'S BOULDER  —  Las Vegas (Off-Strip), NV"); row += 1
    rows = [
        ("  Gaming Revenue ($M)",   28.6,  30.8,  32.8,  34.6,  36.4,  38.2),
        ("  Hotel Revenue ($M)",     4.2,   4.6,   5.0,   5.4,   5.6,   6.0),
        ("  F&B Revenue ($M)",       5.6,   6.0,   6.4,   6.8,   7.2,   7.6),
        ("  Occupancy Rate",        0.652, 0.678, 0.706, 0.728, 0.748, 0.764),
        ("  RevPAR ($)",             14.8,  16.2,  17.8,  19.6,  20.8,  22.0),
        ("  Hotel Rooms",            303,   303,   303,   303,   303,   303),
        ("  Slot Machines",         1061,  1061,  1061,  1061,  1080,  1100),
        ("  Table Games",             16,    16,    16,    16,    16,    16),
        ("  Visitors (000s)",       1820,  1940,  2050,  2160,  2240,  2320),
    ]
    for label, *vals in rows:
        pct = "Rate" in label
        metric(row, label, vals, pct=pct); row += 1

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Debt Schedule
# ═══════════════════════════════════════════════════════════════════════════════
def build_debt_schedule(wb):
    ws = wb.create_sheet("Debt Schedule")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [34, 10, 10, 10, 10, 10, 10])

    YEARS = ['2003A', '2004A', '2005A', '2006E', '2007E', '2008E']

    ws.row_dimensions[1].height = 22
    write_header(ws, 1, "American Casinos  —  Debt Schedule & Credit Statistics ($M)")
    ws.row_dimensions[3].height = 16
    write_col_headers(ws, 3, ["DEBT & LEVERAGE"] + YEARS)

    ws.row_dimensions[4].height = 14
    write_row(ws, 4, "DEBT STRUCTURE", [None]*6, indent=1, bold=True, gray=True)

    debt_rows = [
        ("Senior Secured Notes ($M)",  215.0, 215.0, 215.0, 430.0, 430.0, 430.0),
        ("  Coupon Rate (%)",            8.50,  8.50,  8.50,  8.50,  8.50,  8.50),
        ("  Maturity",              "2012","2012","2012","2012","2012","2012"),
        ("Revolver Capacity ($M)",      40.0,  40.0,  40.0,  60.0,  60.0,  60.0),
        ("Revolver Drawn ($M)",          0.0,   0.0,   0.0,  15.0,   0.0,   0.0),
        ("Total Debt ($M)",            215.0, 215.0, 215.0, 445.0, 430.0, 430.0),
        ("Cash & Equivalents ($M)",     18.4,  24.6,  34.2,  22.8,  38.4,  54.6),
        ("Net Debt ($M)",              196.6, 190.4, 180.8, 422.2, 391.6, 375.4),
    ]
    for i, (label, *vals) in enumerate(debt_rows):
        r = 5 + i
        ws.row_dimensions[r].height = 13
        bold = "Total Debt" in label or "Net Debt" in label
        is_str = isinstance(vals[0], str)
        if is_str:
            lbl = ws.cell(row=r, column=1, value=label)
            lbl.font = Font(name="Calibri", size=9)
            lbl.alignment = Alignment(horizontal="left", indent=2)
            for col, v in enumerate(vals, start=2):
                c = ws.cell(row=r, column=col, value=v)
                c.font = Font(name="Calibri", size=9)
                c.alignment = Alignment(horizontal="center")
        else:
            write_row(ws, r, label, vals, indent=2 if label.startswith("  ") else 1, bold=bold)

    spacer(ws, 13)

    ws.row_dimensions[14].height = 14
    write_row(ws, 14, "CREDIT STATISTICS", [None]*6, indent=1, bold=True, gray=True)

    credit_rows = [
        ("Total Debt / EBITDA (x)",      3.89,  3.13,  2.41,  4.86,  4.24,  3.81),
        ("Net Debt / EBITDA (x)",        3.56,  2.77,  2.02,  4.61,  3.86,  3.32),
        ("EBITDA / Interest (x)",        1.94,  2.41,  3.14,  2.68,  2.97,  3.30),
        ("CapEx / Revenue (%)",         0.051, 0.058, 0.060, 0.066, 0.059, 0.053),
        ("FCF Conversion (%)",          0.227, 0.327, 0.462, 0.316, 0.401, 0.476),
    ]
    for i, (label, *vals) in enumerate(credit_rows):
        r = 15 + i
        ws.row_dimensions[r].height = 13
        pct = "%" in label
        write_row(ws, r, label, vals, indent=2, pct=pct)

    spacer(ws, 20)

    note = ws.cell(row=21, column=1,
        value="Senior Secured Notes issued in two tranches: $215M (2003) + $215M add-on (2006, concurrent with Aquarius acquisition). Revolver matures 2010.")
    note.font = Font(name="Calibri", italic=True, size=8, color="666666")
    note.alignment = Alignment(horizontal="left", indent=1, wrap_text=True)
    ws.merge_cells(start_row=21, start_column=1, end_row=21, end_column=7)
    ws.row_dimensions[21].height = 22

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Transaction Assumptions
# ═══════════════════════════════════════════════════════════════════════════════
def build_transaction(wb):
    ws = wb.create_sheet("Transaction Assumptions")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [34, 14, 10, 10, 10, 10, 10])

    ws.row_dimensions[1].height = 22
    write_header(ws, 1, "American Casinos  —  Transaction Assumptions & Return Analysis")

    def tx_section(r, title):
        ws.row_dimensions[r].height = 15
        write_row(ws, r, title, [None]*6, indent=1, bold=True, gray=True)

    def tx_row(r, label, val, fmt="#,##0.0", pct=False, bold=False):
        ws.row_dimensions[r].height = 13
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = Font(name="Calibri", bold=bold, size=9)
        lbl.alignment = Alignment(horizontal="left", indent=2)
        c = ws.cell(row=r, column=2, value=val)
        c.font = Font(name="Calibri", bold=bold, size=9)
        c.alignment = Alignment(horizontal="right")
        if val is not None:
            c.number_format = '0.0%' if pct else fmt

    # ── Transaction structure ────────────────────────────────────────────────
    tx_section(3, "TRANSACTION STRUCTURE")
    tx_data = [
        ("Enterprise Value ($M)",          950.0, "#,##0.0", False),
        ("EV / LTM EBITDA (x)",              10.6, "#,##0.0x", False),
        ("EV / 2007E EBITDA (x)",             9.4, "#,##0.0x", False),
        ("Equity Purchase Price ($M)",       505.0, "#,##0.0", False),
        ("Assumed Debt at Close ($M)",       430.0, "#,##0.0", False),
        ("Transaction Fees & Expenses ($M)", 14.5, "#,##0.0", False),
        ("Leverage at Close (x)",              4.6, "#,##0.0", False),
        ("Equity Contribution (%)",          0.532, None, True),
    ]
    for i, (label, val, fmt, pct) in enumerate(tx_data):
        r = 4 + i
        tx_row(r, label, val, fmt=fmt or "#,##0.0", pct=pct)

    spacer(ws, 12)

    # ── Return analysis ──────────────────────────────────────────────────────
    tx_section(13, "RETURN ANALYSIS (BASE CASE)")
    ret_data = [
        ("Hold Period (years)",            5,     "#,##0",   False),
        ("Exit EV / EBITDA Multiple (x)",  9.0,   "#,##0.0", False),
        ("Exit Enterprise Value ($M)",     1016.1,"#,##0.0", False),
        ("Less: Exit Debt ($M)",            340.0, "#,##0.0", False),
        ("Exit Equity Value ($M)",          676.1, "#,##0.0", False),
        ("MOIC",                            2.0,   "#,##0.0x",False),
        ("Gross IRR",                       0.148, None,      True),
    ]
    for i, (label, val, fmt, pct) in enumerate(ret_data):
        r = 14 + i
        bold = label in ("Exit Equity Value ($M)", "Gross IRR", "MOIC")
        tx_row(r, label, val, fmt=fmt or "#,##0.0", pct=pct, bold=bold)

    spacer(ws, 21)

    # ── IRR Sensitivity table ────────────────────────────────────────────────
    tx_section(22, "SENSITIVITY  —  IRR vs. EXIT MULTIPLE & EBITDA GROWTH")

    multiples = [7.5, 8.0, 8.5, 9.0, 9.5, 10.0]
    growths   = [0.06, 0.08, 0.10, 0.12, 0.14]

    # Header row
    ws.row_dimensions[23].height = 14
    hdr = ws.cell(row=23, column=1, value="Exit Multiple  \u2192  /  EBITDA Growth  \u2193")
    hdr.font = Font(name="Calibri", bold=True, size=8, color=WHITE)
    hdr.fill = hdr_fill(MID_NAVY)
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for j, m in enumerate(multiples):
        c = ws.cell(row=23, column=2+j, value=f"{m:.1f}x")
        c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        c.fill = hdr_fill(MID_NAVY)
        c.alignment = Alignment(horizontal="center")

    # Sensitivity values — simplified IRR approximations
    irr_table = {
        0.06: [0.098, 0.112, 0.125, 0.138, 0.150, 0.162],
        0.08: [0.110, 0.124, 0.137, 0.149, 0.161, 0.173],
        0.10: [0.121, 0.135, 0.148, 0.160, 0.172, 0.184],
        0.12: [0.132, 0.146, 0.158, 0.170, 0.182, 0.194],
        0.14: [0.143, 0.156, 0.168, 0.181, 0.192, 0.204],
    }
    for i, g in enumerate(growths):
        r = 24 + i
        ws.row_dimensions[r].height = 13
        lbl = ws.cell(row=r, column=1, value=f"{g:.0%} EBITDA CAGR")
        lbl.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        lbl.fill = hdr_fill(MID_NAVY)
        lbl.alignment = Alignment(horizontal="center")
        for j, irr in enumerate(irr_table[g]):
            c = ws.cell(row=r, column=2+j, value=irr)
            c.number_format = '0.0%'
            c.font = Font(name="Calibri", size=9)
            c.alignment = Alignment(horizontal="center")
            # highlight base case ~14.8% IRR
            if abs(irr - 0.148) < 0.008:
                c.fill = hdr_fill(GOLD)
                c.font = Font(name="Calibri", bold=True, size=9, color=DARK_NAVY)
            elif i % 2 == 0:
                c.fill = hdr_fill(LIGHT_GRAY)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove default blank sheet

    build_pl_summary(wb)
    build_property_metrics(wb)
    build_debt_schedule(wb)
    build_transaction(wb)

    out = "C:/Users/hanso/OneDrive/Documents/Coding Projects/sagard-project/cim-analyzer/test/AmericanCasinos_Financial_Model.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
